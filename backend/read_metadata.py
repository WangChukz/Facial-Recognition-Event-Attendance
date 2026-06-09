#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook

wb = load_workbook('./uploads/metadata.xlsx')
ws = wb.active

cols = [cell.value for cell in ws[1]]
print("Columns:", cols)
print("Total rows:", ws.max_row - 1)
print("\nFirst 5 data rows:")

for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
    print(row)
