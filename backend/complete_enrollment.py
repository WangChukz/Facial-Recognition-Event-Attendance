#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Complete enrollment pipeline:
1. Import users from metadata.xlsx
2. Upload card images
3. Extract embeddings (enrollment)
4. Build FAISS index
"""

import asyncio
import sys
import io
from pathlib import Path
from datetime import datetime
import unicodedata



import numpy as np
from openpyxl import load_workbook
from sqlalchemy import select, func

from app.db.session import async_session_maker, engine
from app.db.models import Base, User, UserRole, CardImage, FaceEmbedding
from app.services.face_pipeline import FacePipeline
from app.services.faiss_indexer import FaissFaceIndex
from app.config import get_settings
from app.services.enrollment_v2 import generate_augmented_embeddings


def remove_accents(text: str) -> str:
    """Remove Vietnamese accents."""
    text = text.replace('Đ', 'D').replace('đ', 'd')
    nfd = unicodedata.normalize('NFD', text)
    return ''.join(c for c in nfd if unicodedata.category(c) != 'Mn').replace(' ', '').replace(' ', '')


async def import_users_from_metadata():
    """Import users from metadata.xlsx."""
    print("\n" + "="*70)
    print("STEP 1: Import users from metadata.xlsx")
    print("="*70)

    wb = load_workbook('./uploads/metadata.xlsx')
    ws = wb.active

    users_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        stt, student_code, full_name, lop, dob = row
        email_prefix = remove_accents(str(full_name).strip()).lower()
        users_data.append({
            'student_code': str(student_code).strip(),
            'full_name': str(full_name).strip(),
            'email': f"{email_prefix}@gmail.com",
            'role': UserRole.student,
            'class_name': str(lop).strip() if lop else None,
        })

    print(f"📋 Found {len(users_data)} users in metadata")

    created = 0
    skipped = 0

    async with async_session_maker() as session:
        for user_data in users_data:
            try:
                # Check by student_code (more reliable than email)
                existing = await session.execute(
                    select(User).where(User.student_code == user_data['student_code'])
                )
                if existing.scalar_one_or_none():
                    skipped += 1
                    continue

                user = User(**user_data)
                session.add(user)
                created += 1

                if created % 5 == 0 or created == len(users_data):
                    print(f"  ✅ Created: {user_data['full_name']}")

            except Exception as e:
                await session.rollback()
                print(f"  ⏭️  Skipped: {user_data['full_name']} ({str(e)[:50]}...)")
                skipped += 1
                continue

        try:
            await session.commit()
        except Exception as e:
            await session.rollback()
            print(f"  ⚠️  Commit error: {e}")

    print(f"\n✅ Created: {created} users")
    print(f"⏭️  Already existed: {skipped} users")
    return created + skipped


async def upload_card_images():
    """Upload all card images and save to database."""
    print("\n" + "="*70)
    print("STEP 2: Upload card images")
    print("="*70)

    card_images_dir = Path("./uploads/card_images")
    image_files = sorted(
        list(card_images_dir.glob("*.jpg")) +
        list(card_images_dir.glob("*.jpeg")) +
        list(card_images_dir.glob("*.png")) +
        list(card_images_dir.glob("*.JPG")) +
        list(card_images_dir.glob("*.JPEG")) +
        list(card_images_dir.glob("*.PNG"))
    )

    print(f"📷 Found {len(image_files)} images")

    uploaded = 0
    errors = []

    async with async_session_maker() as session:
        for img_path in image_files:
            try:
                # Parse filename
                name_without_ext = img_path.stem
                parts = name_without_ext.rsplit("_", 1)
                full_name_no_accents = parts[0] if len(parts) >= 1 else name_without_ext
                image_type = parts[1] if len(parts) == 2 else "enroll"

                # Find user
                q = select(User)
                r = await session.execute(q)
                users = r.scalars().all()

                matched_user = None
                for u in users:
                    if remove_accents(u.full_name) == full_name_no_accents:
                        matched_user = u
                        break

                if not matched_user:
                    errors.append((img_path.name, f"User '{full_name_no_accents}' not found"))
                    continue

                # Check if already uploaded
                existing = await session.execute(
                    select(CardImage).where(
                        CardImage.user_id == matched_user.id,
                        CardImage.image_type == image_type
                    )
                )
                if existing.scalar_one_or_none():
                    continue

                # Create CardImage record
                card = CardImage(
                    user_id=matched_user.id,
                    image_type=image_type,
                    image_path=str(img_path),
                    original_filename=img_path.name,
                )
                session.add(card)
                uploaded += 1

                if uploaded % 5 == 0 or uploaded == len(image_files):
                    print(f"  ✅ {matched_user.full_name}")

            except Exception as e:
                errors.append((img_path.name, str(e)))

        await session.commit()

    print(f"\n✅ Uploaded: {uploaded} images")
    if errors:
        print(f"❌ Errors: {len(errors)}")
        for fname, err in errors[:3]:
            print(f"   - {fname}: {err}")


async def enrollment_all_cards():
    """Extract embeddings from all card images."""
    print("\n" + "="*70)
    print("STEP 3: Enrollment - Extract embeddings from card images")
    print("="*70)

    settings = get_settings()
    pipeline = FacePipeline()
    faiss_index = FaissFaceIndex(settings.faiss_index_path, settings.faiss_meta_path)
    faiss_index.load()

    enrolled = 0
    errors = []

    # Get all card images
    async with async_session_maker() as session:
        r = await session.execute(select(CardImage))
        cards = r.scalars().all()

    print(f"📸 Processing {len(cards)} card images...")
    import sys

    for card in cards:
        try:
            # Check if this user already has embeddings
            async with async_session_maker() as session:
                existing_emb = await session.execute(
                    select(FaceEmbedding).where(FaceEmbedding.user_id == card.user_id)
                )
                if existing_emb.scalars().all():
                    print(f"  ⏭️  Skipping {card.original_filename} (already has embeddings)")
                    sys.stdout.flush()
                    continue

            # Read image
            if not Path(card.image_path).exists():
                errors.append((card.original_filename, "File not found"))
                continue

            with open(card.image_path, 'rb') as f:
                img_data = f.read()

            img = pipeline.decode_image_bytes(img_data)

            # Process frame - detect face and extract embedding
            faces = pipeline.process_frame_sync(img, use_adaptive_clahe=True)

            if not faces:
                errors.append((card.original_filename, "No face detected"))
                continue

            # Use best face
            best_face = max(faces, key=lambda x: x['det_score'])

            if best_face['det_score'] < 0.50:
                errors.append((card.original_filename, f"Low confidence: {best_face['det_score']}"))
                continue

            # Extract augmented embeddings sequentially
            def process_for_augment(frame):
                results = pipeline.process_frame_sync(frame, use_adaptive_clahe=True)
                return results if results else []

            embeddings = generate_augmented_embeddings(
                img,
                process_for_augment,
                n_geometric=7,
                n_photo=5,
                n_combined=2,
                n_occlusion=2,
            )

            if not embeddings:
                errors.append((card.original_filename, "Could not extract augmented embeddings"))
                continue

            from app.services.attendance_logic import next_faiss_id
            
            # Store all augmented embeddings in DB and FAISS
            async with async_session_maker() as session:
                for i, emb in enumerate(embeddings):
                    fid = await next_faiss_id(session)
                    emb_bytes = np.asarray(emb, dtype=np.float32).tobytes()
                    fe = FaceEmbedding(
                        user_id=card.user_id,
                        faiss_id=fid,
                        embedding_dim=len(emb),
                        embedding_vector=emb_bytes,
                        is_primary=(i == 0),
                        image_path=f"{card.image_path}_aug{i:02d}" if i > 0 else card.image_path,
                    )
                    session.add(fe)
                    await session.flush()
                    faiss_index.add_with_id(emb, fid, fe.id, card.user_id)
                
                await session.commit()
            
            enrolled += 1

            # Print progress and flush
            async with async_session_maker() as session:
                user = await session.get(User, card.user_id)
                user_name = user.full_name if user else "Unknown"
            print(f"  ✅ {user_name} (confidence: {best_face['det_score']:.2f}, embeddings: {len(embeddings)})")
            sys.stdout.flush()

        except Exception as e:
            print(f"  ❌ Error: {card.original_filename}: {e}")
            sys.stdout.flush()
            errors.append((card.original_filename, str(e)))

    # Persist FAISS index
    print("\n💾 Persisting FAISS index...")
    sys.stdout.flush()
    faiss_index.persist()

    print(f"\n✅ Enrolled: {enrolled} users with augmented embeddings")
    if errors:
        print(f"❌ Errors: {len(errors)}")
        for fname, err in errors[:3]:
            print(f"   - {fname}: {err}")
    sys.stdout.flush()

    return enrolled


async def verify_enrollment():
    """Verify enrollment data."""
    print("\n" + "="*70)
    print("STEP 4: Verify enrollment data")
    print("="*70)

    async with async_session_maker() as session:
        # Count users
        r = await session.execute(select(func.count()).select_from(User))
        user_count = r.scalar()

        # Count card images
        r = await session.execute(select(func.count()).select_from(CardImage))
        card_count = r.scalar()

        # Count embeddings
        r = await session.execute(select(func.count()).select_from(FaceEmbedding))
        emb_count = r.scalar()

    print(f"📊 Statistics:")
    print(f"   👥 Users: {user_count}")
    print(f"   📷 Card Images: {card_count}")
    print(f"   🧠 Face Embeddings: {emb_count}")
    print(f"\n✅ System ready for testing!")


async def main():
    """Run complete enrollment pipeline."""
    print("\n" + "="*70)
    print("🚀 COMPLETE ENROLLMENT PIPELINE")
    print("="*70)

    try:
        # Create tables
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)
        print("✅ Database tables created")

        # Step 1: Import users
        await import_users_from_metadata()

        # Step 2: Upload images
        await upload_card_images()

        # Step 3: Enrollment
        await enrollment_all_cards()

        # Step 4: Verify
        await verify_enrollment()

        print("\n" + "="*70)
        print("✅ ENROLLMENT COMPLETE!")
        print("="*70)
        print("\n📝 Next steps:")
        print("   1. Start server: python -m uvicorn app.main:app --reload --port 8000")
        print("   2. Test recognition: Open web app and test face recognition")
        print("\n")

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        await engine.dispose()


if __name__ == "__main__":
    asyncio.run(main())
