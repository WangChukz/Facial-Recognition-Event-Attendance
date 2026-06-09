#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Import users from metadata.xlsx to database."""

import asyncio
import sys
import io
from datetime import datetime
from pathlib import Path
import unicodedata

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook
from app.db.session import async_session_maker
from app.db.models import User, UserRole


def remove_accents(text):
    """Remove Vietnamese accents: Bùi Đức Thịnh → Bui Duc Thinh."""
    nfd = unicodedata.normalize('NFD', text)
    return ''.join(c for c in nfd if unicodedata.category(c) != 'Mn')


async def import_users():
    """Import users from metadata.xlsx."""
    wb = load_workbook('./uploads/metadata.xlsx')
    ws = wb.active

    users_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        stt, student_code, full_name, lop, dob = row
        users_data.append({
            'student_code': str(student_code),
            'full_name': str(full_name),
            'email': f"{student_code}@student.edu.vn",
            'role': UserRole.student,
        })

    print(f"📋 Found {len(users_data)} users in metadata")
    print(f"\n📝 Creating users...")

    created = 0
    skipped = 0
    errors = []

    async with async_session_maker() as session:
        for user_data in users_data:
            try:
                # Check if exists
                from sqlalchemy import select
                existing = await session.execute(
                    select(User).where(User.email == user_data['email'])
                )
                if existing.scalar_one_or_none():
                    skipped += 1
                    continue

                user = User(**user_data)
                session.add(user)
                created += 1

                # Print progress
                if created % 5 == 0:
                    print(f"  ✅ {user_data['full_name']}")

            except Exception as e:
                errors.append((user_data['full_name'], str(e)))

        await session.commit()

    print(f"\n✅ Created: {created}")
    print(f"⏭️  Skipped: {skipped}")
    if errors:
        print(f"❌ Errors: {len(errors)}")
        for name, err in errors[:3]:
            print(f"   - {name}: {err}")


if __name__ == "__main__":
    asyncio.run(import_users())
